#!/usr/bin/env python3
"""Fetch TKO Sports Ground timetable from LCSD XLSX → write sport_ground_status.json.
Usage: python3 sport_ground_sync.py [--deploy]
Parses Main Field + Secondary Field timetables (A/L/B/M codes) + Notices sheet.
"""
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.request
from datetime import date, datetime, timezone, timedelta

FID = 1060
BASE_URL = "https://www.lcsd.gov.hk/file_upload_clpss/leisure_facilities/lsb/jogging"
JSON_PATH = "/home/ubuntu/we1co.me/sport_ground_status.json"
UA = {"User-Agent": "Mozilla/5.0 (compatible; sport-ground-sync/1.0)"}
STATUS_MAP = {
    "A": {"en": "Open", "zh": "開放"},
    "L": {"en": "Limited lanes", "zh": "部分線道"},
    "B": {"en": "Closed (booking)", "zh": "預訂暫停"},
    "M": {"en": "Closed", "zh": "關閉"},
}
NOTICE_COL = {"B": "date_range", "C": "facilities", "D": "reason", "E": "remarks"}


def fetch_xlsx(month=None):
    """Download XLSX for given month via CF proxy, fallback to prev/next.
    Primary call (month=None): always try current month first — it's today's data.
    Secondary call (month=specific): try that month first — it's tomorrow's data."""
    now = datetime.now(timezone(timedelta(hours=8)))
    if month is None:
        month = now.strftime("%Y%m")
    prev = (now.replace(day=1) - timedelta(days=1)).strftime("%Y%m")
    nxt = (now.replace(day=1) + timedelta(days=32)).replace(day=1).strftime("%Y%m")
    # Always try requested month first, then prev, then nxt
    urls = [f"{BASE_URL}/{FID}_{month}.xlsx", f"{BASE_URL}/{FID}_{prev}.xlsx", f"{BASE_URL}/{FID}_{nxt}.xlsx"]
    # Try CF proxy first
    try:
        import urllib.parse as urlparse
        token = None
        for p in ["/home/ubuntu/.config/lcsd_proxy_token", "/tmp/cf_proxy_token.txt"]:
            try:
                with open(p) as f: token = f.read().strip(); break
            except (FileNotFoundError, PermissionError): pass
        if not token:
            token = os.environ.get("CF_PROXY_TOKEN", "")
        proxy_base = "https://lcsd-proxy.forumdata.workers.dev/?url="
        if token:
            for url in urls:
                try:
                    proxied = proxy_base + urlparse.quote(url, safe='')
                    req = urllib.request.Request(proxied, headers={**UA, "x-proxy-token": token})
                    with urllib.request.urlopen(req, timeout=30) as r:
                        data = r.read()
                        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                            f.write(data)
                            print(f"Fetched via CF proxy: {url}")
                            return f.name
                except Exception as e:
                    print(f"  skip CF {url}: {e}")
    except Exception:
        pass
    for url in urls:
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read()
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                    f.write(data)
                    print(f"Fetched: {url}")
                    return f.name
        except Exception as e:
            print(f"  skip {url}: {e}")
    raise RuntimeError("Cannot fetch XLSX")


def parse_timetable(ws, today_day=None):
    """Parse timetable sheet: row 8=dates, row 9+=time slots × status codes.
    Returns dict: {time_slot: {day_num: code}} + list of date headers.
    """
    dates = {}  # col_letter -> day_num
    timetable = {}  # time_slot -> {day_num: code}

    # Row 8: date headers
    for cell in ws[8]:
        if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
            lines = cell.value.strip().split("\n")
            try:
                day_num = int(lines[0])
            except ValueError:
                continue
            dates[cell.column_letter] = day_num

    # Row 9+: time slots (col B = time, col C+ = status codes)
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=False):
        # row[0] is col A (empty), row[1] is col B (time)
        time_cell = row[1] if len(row) > 1 else row[0]
        if not time_cell.value:
            continue
        time_slot = str(time_cell.value).strip()
        codes = {}
        for cell in row[2:]:
            if cell.column_letter in dates and cell.value:
                code = str(cell.value).strip().upper()
                if code in STATUS_MAP:
                    codes[dates[cell.column_letter]] = code
        if codes:
            timetable[time_slot] = codes

    return timetable, dates


def parse_notices(ws):
    """Parse Notices sheet for closure information."""
    notices = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        vals = {}
        for cell in row:
            if cell.value and cell.column_letter in NOTICE_COL:
                vals[NOTICE_COL[cell.column_letter]] = str(cell.value).strip()
        if vals.get("date_range") and vals.get("facilities"):
            notices.append(vals)
    return notices


def get_current_status(timetable, dates, notices, day_num, field_type="main"):
    """Get status for each time slot for given day based on A/L/B/M codes.
    The XLSX codes are authoritative — they already encode closures/maintenance.
    Notices are returned separately for display only (not applied to codes)."""
    now_hkt = datetime.now(timezone(timedelta(hours=8)))
    current_minutes = now_hkt.hour * 60 + now_hkt.minute

    # Build list of notice closures for this day with parsed time ranges
    today_str = date.today().strftime("%Y/%m/%d")
    notice_closures = []  # [(start_min, end_min, facilities_str, reason_str)]
    for n in notices:
        dr = n.get("date_range", "")
        date_part = dr.split("  ")[0] if "  " in dr else dr
        time_part = dr.split("  ")[1] if "  " in dr else ""
        # Parse time range from notice (e.g. "06:30-17:00" or "06:30-22:30 / 06:30-20:30")
        tm = re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})", time_part)
        if not tm:
            continue
        n_start = int(tm.group(1)[:2]) * 60 + int(tm.group(1)[3:5])
        n_end = int(tm.group(2)[:2]) * 60 + int(tm.group(2)[3:5])
        fac = n.get("facilities", "")
        reason = n.get("reason", "")
        # Filter by field type: only apply notices relevant to this field
        fac_lower = fac.lower()
        if field_type == "main" and "主場" not in fac and "main field" not in fac_lower:
            continue  # skip notices that don't mention main field
        elif field_type == "secondary" and "副場" not in fac and "secondary" not in fac_lower:
            continue  # skip notices that don't mention secondary field
        # Check if this notice matches day_num
        date_section = date_part.split(",")[0] if "," in date_part else date_part
        m = re.match(r"(\d{4}/\d{2}/\d{2})", date_section)
        if m:
            base_date = m.group(1)
            parts = date_section.split(",")
            for p in parts:
                p = p.strip()
                if re.match(r"\d{4}/\d{2}/\d{2}", p):
                    # Full date — check if it equals today AND day matches
                    # We need to match by day_num, not just today_str
                    pass  # handled below via day_num comparison
                elif re.match(r"\d{1,2}$", p):
                    pass  # handled below
            # Simpler: check if day_num appears in the notice's day list
            # Extract all day numbers from the notice
            all_days = set()
            for p in date_section.split(","):
                p = p.strip().split("\n")[0].strip()
                dm = re.match(r"\d{4}/\d{2}/(\d{2})", p)
                if dm:
                    all_days.add(int(dm.group(1)))
                elif re.match(r"\d{1,2}$", p):
                    all_days.add(int(p))
                # Handle ranges like "12-13"
                rm = re.match(r"(\d{4}/\d{2}/(\d{2}))-(\d{2})", p)
                if rm:
                    for d in range(int(rm.group(2)), int(rm.group(3)) + 1):
                        all_days.add(d)
            if day_num in all_days:
                notice_closures.append((n_start, n_end, fac, reason))

    result = []
    for time_slot, day_codes in timetable.items():
        time_slot_clean = time_slot.strip()
        m = re.match(r"(\d{2}):(\d{2})\s*-\s*(\d{2}):(\d{2})", time_slot_clean)
        if not m:
            continue
        sh, sm, eh, em = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        start_min = sh * 60 + sm
        end_min = eh * 60 + em

        code = day_codes.get(day_num, "")
        status = STATUS_MAP.get(code, {"en": "Unknown", "zh": "未知"})

        is_current = start_min <= current_minutes < end_min

        result.append({
            "time": time_slot,
            "code": code,
            "status_zh": status["zh"],
            "status_en": status["en"],
            "is_current": is_current,
        })

    # Build closure list for display
    closures = []
    for n_start, n_end, fac, reason in notice_closures:
        closures.append({
            "date": today_str,
            "facilities": fac,
            "reason": reason,
            "time": f"{n_start//60:02d}:{n_start%60:02d}-{n_end//60:02d}:{n_end%60:02d}",
        })

    return result, closures


def _dedupe_closures(closures):
    """Remove duplicate closure notices (e.g. 主場及副場 appearing for both fields)."""
    seen, out = set(), []
    for c in closures:
        k = (c.get("date", ""), c.get("time", ""), c.get("facilities", "").strip())
        if k in seen:
            continue
        seen.add(k)
        out.append(c)
    return out


def _notice_weekdays(closure):
    """Return weekday numbers (0=Mon..6=Sun) covered by a 場地保養 closure notice.
    Only maintenance notices (reason contains 保養) count."""
    reason = closure.get("reason", "")
    if "保養" not in reason:
        return set()
    d = closure.get("date", "")
    if not d:
        return set()
    try:
        dt = datetime.strptime(d, "%Y/%m/%d")
        return {dt.weekday()}
    except (ValueError, TypeError):
        return set()


def _maint_summary(notices, field):
    """Human-readable maintenance summary from 場地保養 notices, e.g. '逢星期一'."""
    wd_cn = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    wds = set()
    for n in notices:
        reason = n.get("reason", "")
        fac = n.get("facilities", "")
        if "保養" not in reason:
            continue
        if field == "main" and "主場" not in fac and "main field" not in fac.lower():
            continue
        if field == "secondary" and "副場" not in fac and "secondary" not in fac.lower():
            continue
        dr = n.get("date_range", "")
        date_part = dr.split("  ")[0] if "  " in dr else dr
        # weekday of the first date in the notice
        m = re.search(r"(\d{4}/\d{2}/\d{2})", date_part)
        if m:
            try:
                dt = datetime.strptime(m.group(1), "%Y/%m/%d")
                wds.add(dt.weekday())
            except ValueError:
                pass
    if not wds:
        return ""
    if len(wds) == 1:
        return f"逢{wd_cn[min(wds)]}"
    return "、".join(f"逢{wd_cn[w]}" for w in sorted(wds))


def main():
    try:
        xlsx_path = fetch_xlsx()
        try:
            import openpyxl
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        today_day = date.today().day

        # Also try next month XLSX for cross-month tomorrow
        now_hkt = datetime.now(timezone(timedelta(hours=8)))
        tomorrow = now_hkt + timedelta(days=1)
        cross_month = tomorrow.month != now_hkt.month
        wb2 = None
        if cross_month:
            try:
                tomorrow_month = tomorrow.strftime("%Y%m")
                nxt_path = fetch_xlsx(tomorrow_month)
                if nxt_path: wb2 = openpyxl.load_workbook(nxt_path)
            except (FileNotFoundError, PermissionError): pass

        # Parse both sheets
        main_ws = wb[wb.sheetnames[0]]
        sec_ws = wb[wb.sheetnames[1]]
        notice_ws = wb[wb.sheetnames[2]] if len(wb.sheetnames) > 2 else None

        main_timetable, main_dates = parse_timetable(main_ws)
        sec_timetable, sec_dates = parse_timetable(sec_ws)
        # Merge next month timetable (cross-month tomorrow)
        if wb2:
            for nm_ws in [wb2[wb2.sheetnames[0]], wb2[wb2.sheetnames[1]] if len(wb2.sheetnames)>1 else None]:
                if not nm_ws: continue
                nm_tt, nm_dates = parse_timetable(nm_ws)
                # nm_dates maps col→day_num (1,2,3... for next month)
                # We want day_num 1 (tomorrow=Sep 1) → map to our day 1
                for ts, codes in nm_tt.items():
                    if ts in (main_timetable if nm_ws == wb2[wb2.sheetnames[0]] else sec_timetable):
                        target = main_timetable if nm_ws == wb2[wb2.sheetnames[0]] else sec_timetable
                        for dn, code in codes.items():
                            target.setdefault(ts, {})[dn] = code
        notices = parse_notices(notice_ws) if notice_ws else []

        # Extract XLSX date from sheet title (e.g. "2026年8月")
        title_cell = main_ws['B1']
        xlsx_date = ""
        if title_cell.value:
            dm = re.search(r"(\d{4})年(\d{1,2})月", str(title_cell.value))
            if dm:
                xlsx_date = f"{dm.group(1)}年{dm.group(2)}月"

        # Extract file modified date from XLSX internal metadata (docProps/core.xml)
        xlsx_file_date = ""
        try:
            import zipfile
            with zipfile.ZipFile(xlsx_path) as zf:
                if 'docProps/core.xml' in zf.namelist():
                    core_xml = zf.read('docProps/core.xml').decode('utf-8', errors='ignore')
                    # Look for dcterms:modified or created
                    mod_match = re.search(r'<dcterms:modified[^>]*>([^<]+)</dcterms:modified>', core_xml)
                    if mod_match:
                        # Parse ISO format like 2026-08-29T10:30:00Z
                        raw = mod_match.group(1).strip()
                        dt_match = re.match(r'(\d{4})-(\d{2})-(\d{2})', raw)
                        if dt_match:
                            xlsx_file_date = f"{dt_match.group(1)}-{dt_match.group(2)}-{dt_match.group(3)}"
        except Exception:
            pass

        tomorrow_day = today_day + 1
        # Check if tomorrow exceeds current month's days
        import calendar
        now_hkt = datetime.now(timezone(timedelta(hours=8)))
        days_in_month = calendar.monthrange(now_hkt.year, now_hkt.month)[1]
        if tomorrow_day > days_in_month:
            tomorrow_day = 1  # wrap to next month (day 1)

        # Today
        main_status, main_closures = get_current_status(main_timetable, main_dates, notices, today_day, "main")
        sec_status, sec_closures = get_current_status(sec_timetable, sec_dates, notices, today_day, "secondary")

        # Tomorrow
        main_status_tm, _ = get_current_status(main_timetable, main_dates, notices, tomorrow_day, "main")
        sec_status_tm, _ = get_current_status(sec_timetable, sec_dates, notices, tomorrow_day, "secondary")

        # Maintenance overrides — XLSX is authoritative (already has M codes).
        # Only fill M into slots WITHOUT a valid code, and only on days where a
        # 場地保養 notice exists — never blanket by weekday (maintenance days vary).
        _wd = now_hkt.weekday()
        maint_days_main = {_d for _c in main_closures for _d in _notice_weekdays(_c)}
        maint_days_sec = {_d for _c in sec_closures for _d in _notice_weekdays(_c)}
        is_main_maint = _wd in maint_days_main
        is_sec_maint = _wd in maint_days_sec
        if is_main_maint:
            for s in main_status:
                if s["code"] not in ("A", "B", "L"):
                    s["code"] = "M"
                    s["status_zh"] = "關閉"
                    s["status_en"] = "Closed"
                    s["is_current"] = False
        if is_sec_maint:
            for s in sec_status:
                if s["code"] not in ("A", "B", "L"):
                    s["code"] = "M"
                    s["status_zh"] = "關閉"
                    s["status_en"] = "Closed"
                    s["is_current"] = False
        # Also check tomorrow's maintenance for tomorrow slots
        _wd_tom = (now_hkt + timedelta(days=1)).weekday()
        if _wd_tom in maint_days_main:
            for s in main_status_tm:
                if s["code"] not in ("A", "B", "L"):
                    s["code"] = "M"
                    s["status_zh"] = "關閉"
                    s["status_en"] = "Closed"
        if _wd_tom in maint_days_sec:
            for s in sec_status_tm:
                if s["code"] not in ("A", "B", "L"):
                    s["code"] = "M"
                    s["status_zh"] = "關閉"
                    s["status_en"] = "Closed"
        # Overall: maintenance forces closed for that field, mixed → partial
        all_codes_main = set(d.get(today_day, "") for d in main_timetable.values())
        all_codes_sec = set(d.get(today_day, "") for d in sec_timetable.values())
        # Override with maintenance
        if is_main_maint:
            all_codes_main = {"M"}
        if is_sec_maint:
            all_codes_sec = {"M"}
        all_codes = all_codes_main | all_codes_sec
        all_codes.discard("")
        if is_main_maint and is_sec_maint:
            overall = "closed"
        elif is_main_maint or is_sec_maint:
            overall = "partial"
        elif not all_codes:
            overall = "unknown"
        elif "A" in all_codes or "L" in all_codes:
            overall = "partial" if len(all_codes) > 1 else "open"
        elif "B" in all_codes or "M" in all_codes:
            overall = "closed"
        else:
            overall = "unknown"

        sync_time = datetime.now(now_hkt.tzinfo).strftime("%Y-%m-%d %H:%M")

        result = {
            "lastSync": sync_time,
            "today": today_day,
            "todayDate": date.today().strftime("%Y/%m/%d"),
            "tomorrow": tomorrow_day,
            "tomorrowDate": (now_hkt + timedelta(days=1)).strftime("%Y/%m/%d"),
            "overall": overall,
            "mainField": main_status,
            "secondaryField": sec_status,
            "mainFieldTomorrow": main_status_tm,
            "secondaryFieldTomorrow": sec_status_tm,
            "closures": _dedupe_closures(main_closures + sec_closures),
            "maintenance": {"main": _maint_summary(notices, "main"), "sec": _maint_summary(notices, "secondary")},
            "noticeUrl": f"https://www.lcsd.gov.hk/clpss/tc/webApp/Facility/Details.do?fid={FID}",
            "xlsxUrl": f"{BASE_URL}/{FID}_{(now_hkt + timedelta(days=1)).strftime('%Y%m') if tomorrow_day == 1 else now_hkt.strftime('%Y%m')}.xlsx",
            "xlsxDate": (lambda d: (lambda m: f"{m[1]}年{str(int(m[2])+1).zfill(2)}月" if int(m[2])<12 else f"{int(m[1])+1}年01月")(re.match(r"(\d{4})年(\d{1,2})月", d)))(xlsx_date) if tomorrow_day == 1 and re.match(r"\d{4}年\d{1,2}月", xlsx_date) else xlsx_date,
            "xlsxFileDate": xlsx_file_date,
        }

        with open(JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        os.unlink(xlsx_path)
        print(f"sport_ground_status.json: {overall}, {len(main_status)} main slots, {len(sec_status)} sec slots, {len(notices)} notices, synced {sync_time}")
        print("PASS")

    except Exception as e:
        print(f"FAIL: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
